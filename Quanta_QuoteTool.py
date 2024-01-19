"""
Update by cory.chang@hp.com for Python3.12 - Quanta Quote tool
version: 01/19/2024
"""

from openpyxl import load_workbook
import pandas as pd
import numpy as np
import pyxlsb 
import os
import decimal
import xlwings as xw
import datetime
import re
from PyQt5.QtWidgets import *

"""
This Progarm was built by Lily(lily.chen1@hp.com) for quote validation - Quanta
version: 07/07/2022
"""
## 1.function1: Delete CPCT non-effectively date ---Start-------------------------------------------------------------------------------------------
def cpct_checker():
    ## -----below code is for testing
    # simulation = 'C:/Users/CheLily/OneDrive - HP Inc/Desktop/Quote/Quote Tool/test_iec/Quote Tool_20220303.xlsm' #for test
    # wb = xw.Book(simulation) #for test
    # base = wb.sheets['Main'] #for test
    ## -----
    base = xw.Book.caller().sheets['Main'] 
    base.range('C43', 'L87').clear_contents()
    path_cpct = base.range('F4').value  
    validated_date = base.range('F8').value 
    all_paths_cpct = []
    all_names_cpct = []
    for dirPath, dirNames, fileNames in os.walk(path_cpct):
        for f in fileNames:
            all_paths_cpct.append(os.path.join(dirPath, f))  
            all_names_cpct.append(f)

    if isinstance(validated_date,datetime.datetime) == False:
        formatted_date1 = datetime.datetime.strptime(validated_date, "%m/%Y")
    else:
        formatted_date1 = validated_date
        
    row_no = 43
    if_tbd = 0
    for file, name in zip(all_paths_cpct, all_names_cpct) :
        revised_column_no = 0
        revised_column_no2 = 0
        #get sheet's name
        cpc_traker = load_workbook(file)
        sheet_names = cpc_traker.get_sheet_names()
        pricing_update = []
        spec_change = []
        for i in sheet_names:
            if 'CPC Tracker-pricing update' in i:
                pricing_update = i
            elif 'CPC Tracker-Pricing update' in i:
                pricing_update = i
            elif 'CPC Tracker-Pricing Update' in i:
                pricing_update = i
            if 'CPC Tracker-spec change' in i:
                spec_change = i
            elif 'CPC Tracker-Spec change' in i:
                spec_change = i
            elif 'CPC Tracker-Spec Change' in i:
                spec_change = i
        #load data
        worksheet1 = cpc_traker[pricing_update]
        if worksheet1['A3'].value == None:
            df1 = pd.read_excel(file, skiprows = 3, sheet_name = pricing_update)
            df2 = pd.read_excel(file, skiprows = 3, sheet_name = spec_change)
        else:
            df1 = pd.read_excel(file, skiprows = 2, sheet_name = pricing_update)
            df2 = pd.read_excel(file, skiprows = 2, sheet_name = spec_change) 
    
  


    #=============================================find 'CPC Tracker-pricing update' should be cleared place
        
        #find row and column of the cell that should be blank
        t = df1.loc[:,'Requestor':'Description'].head()
        revised_columns = t.iloc[:,1:-1].columns


        revised_index = []
        tbd_index = []
        for i, j, z in zip(df1['Effectivity Date'], df1['Request Date'], df1.index):
            #get the index is not effective
            
            #1.  Effectivity Date > validate date
            if isinstance(i,datetime.datetime) == True:                  
                if i > formatted_date1:
                    revised_index.append(z)
            #2.  Effectivity Date = Imme, and  Request Date > validate date
            elif i == 'Immed':
                if j > formatted_date1:
                    revised_index.append(z)
            #3.  TBD
            elif i == 'TBD':
                    tbd_index.append(z)         

        #get the index of the cell that should be blank
        revised_column_index = []
        for i in revised_columns.tolist():
            revised_column_index.append(df1.columns.get_loc(i))


        #=============================================find 'CPC Tracker-spec change' should be cleared place
        #find row and column of the cell that should be blank
        t2 = df2.loc[:,'Requestor':'RFQ (Before)'].head()
        revised_columns2 = t2.iloc[:,1:-1].columns

        revised_index2 = []
        tbd_index2 = []
        for i, j, z in zip(df2['Effectivity Date'], df2['Request Date'], df2.index):
            #get the index is not effective
            
            #1.  Effectivity Date > validate date
            if isinstance(i,datetime.datetime) == True:                  
                if i > formatted_date1:
                    revised_index2.append(z)
            #2.  Effectivity Date = Imme, and  Request Date > validate date
            elif i == 'Immed':
                if j > formatted_date1:
                    revised_index2.append(z)
            #3.  TBD
            elif i == 'TBD':
                    tbd_index2.append(z)

        
        #get the index of the cell that should be blank
        revised_column_index2 = []
        for i in revised_columns2.tolist():
            revised_column_index2.append(df2.columns.get_loc(i))            

        #=============================================CPC Tracker-pricing update    
        #color the cell that is not effective
        #app = xw.App(visible = False)
        #app.display_alerts = False
        #wb2 = xw.Book(file)
        app = xw.App(visible=False, add_book=False) # can't see the file
        app.display_alerts = False # close alert
        app.screen_updating = False # close screen update
        wb2 = app.books.open(file, update_links=False)
        
        sheet1 = wb2.sheets[pricing_update]
        
        for i in revised_column_index: #column
            for j in revised_index: #row
                if sheet1.range(j+4,i+1).value != np.nan and sheet1.range(j+4,i+1).value != None:
                    sheet1.range(j+4,i+1).color = (192, 0, 0)
                    sheet1.range(j+4,i+1).value = np.nan
                    revised_column_no +=1
        
        for i in revised_column_index: #column
            for j in tbd_index: #row
                if sheet1.range(j+4,i+1).value != np.nan and sheet1.range(j+4,i+1).value != None:
                    sheet1.range(j+4,i+1).color = (192, 0, 0)
                    sheet1.range(j+4,i+1).value = np.nan
                    if_tbd = 1
                    print_tbd = 1
                    
        #=============================================CPC Tracker-spec change

        #color the cell that is not effective
        sheet2 = wb2.sheets[spec_change]
        for i in revised_column_index2: #column
            for j in revised_index2: #row
                if sheet2.range(j+4,i+1).value != np.nan and sheet2.range(j+4,i+1).value != None:
                    sheet2.range(j+4,i+1).color = (192, 0, 0)
                    sheet2.range(j+4,i+1).value = np.nan
                    revised_column_no2 +=1
        for i in revised_column_index2: #column
            for j in tbd_index2: #row
                if sheet2.range(j+4,i+1).value != np.nan and sheet2.range(j+4,i+1).value != None:
                    sheet2.range(j+4,i+1).color = (192, 0, 0)
                    sheet2.range(j+4,i+1).value = np.nan
                    if_tbd = 1
                    print_tbd = 1
                    
        #show in the sheets 
        #  base_cpct = wb.sheets['CPCT'] #for test
        base_cpct = xw.Book.caller().sheets['Main'] 

        base_cpct.range(row_no,3).value = name
        base_cpct.range(row_no,6).value = revised_column_no   
        base_cpct.range(row_no,9).value = revised_column_no2
        if if_tbd == 1:
            base_cpct.range(row_no,12).value = "Please check your CPCT file: TBD with subcategory"
            print_tbd = 0
        row_no += 1

        # Create target Directory if it doesn't exist
        dirName = path_cpct+'_revised'
        if not os.path.exists(dirName):
            os.mkdir(dirName) 
    
        wb2.save(dirName + '/' + name)
        wb2.close()
        app.kill()
    if if_tbd == 1:
        tbd_warn_code()
    finish_code()

## 1-1.function1-1: Let user user know the code is finished (pop up Done buttion.) ---Start-----------------------------------------       

def finish_code():
    app = QApplication(sys.argv)
    closeForm = CloseForm('Quote Validation')
    closeForm.show()
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...')
        
class TBD_Warn_Form(QWidget):
    def __init__(self, name = 'TBD_Warn_Form'):
        super(TBD_Warn_Form,self).__init__()
        self.setWindowTitle(name)
        self.resize(300,150)   # set the pop up widget's size

        # btn 1
        self.btn_done = QPushButton(self)  
        self.btn_done.setObjectName("btn_tbd_warn")  
        self.btn_done.setText("Please check your CPCT file: TBD with subcategory")
        
        # set the widget's layout
        layout = QVBoxLayout()
        layout.addWidget(self.btn_done)


        self.setLayout(layout)


        # set the widget's signal
        self.btn_done.clicked.connect(self.close)
        
def tbd_warn_code():
    app = QApplication(sys.argv)
    tbd_warn_Form = TBD_Warn_Form('Quote Validation')
    tbd_warn_Form.show()
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...')
## 1-1.function1-1: Let user user know the code is finished (pop up Done buttion.) ---Done-----------------------------------------       
        

    
## 1.function1: Delete CPCT non-effectively date ---End-------------------------------------------------------------------------------------------

## 2.function2: Quote Validation ---Start-----------------------------------------------------------------------------------------------------------------------------------------------------------------
def Quote_Validation():
    base = xw.Book.caller().sheets['Main'] 
    paths_quote = base.range('F13').value  
    paths_cpct = base.range('F17').value  
    paths_ckit_1 = base.range('G21').value  
    paths_ckit_2 = base.range('G22').value 
    paths_ckit_3 = base.range('G23').value 
    sell_pricing = base.range('F25').value  
    cost_matrix = base.range('F29').value  
    kbd = base.range('F33').value
    
    #get quote_program_matrix sheet's name
    if 'xlsb' in paths_quote:
        quote_program_matrix = pd.ExcelFile(paths_quote, engine='pyxlsb')
    else:
        quote_program_matrix = pd.ExcelFile(paths_quote)
    
    sheet_names = quote_program_matrix.sheet_names 
    quote_program_matrix_update = []
    spec_change = []
    sheet_name_program_matrix = []
    sheet_name_ckit = []
    sheet_name_masterdata = []
    sheet_name_busa = []
    sheet_name_srpbom = []
    sheet_name_avsummary = []
    sheet_name_skusummary = []
    for i in sheet_names:
        if 'Program Matrix' in i:
            sheet_name_program_matrix = i
        if 'CKIT' in i: 
            sheet_name_ckit = i
        if 'Master Data' in i: 
            sheet_name_masterdata = i
        if 'BU SA' in i: 
            sheet_name_busa = i
        if 'BUSA' in i: 
            sheet_name_busa = i
        if 'SRP BOM' in i: 
            sheet_name_srpbom = i
        if 'AV Summary' in i: 
            sheet_name_avsummary = i
        if 'SKU Summary' in i: 
            sheet_name_skusummary = i
    
    #load the data
    if 'xlsb' in paths_quote:
        df_quote_programmatrix = pd.read_excel(paths_quote, engine='pyxlsb', skiprows = 4, sheet_name = sheet_name_program_matrix)    
        df_quote_srpbom = pd.read_excel(paths_quote, engine='pyxlsb', skiprows = 0, sheet_name = sheet_name_srpbom)
    else: 
        df_quote_programmatrix = pd.read_excel(paths_quote, skiprows = 4, sheet_name = sheet_name_program_matrix)
        df_quote_srpbom = pd.read_excel(paths_quote, skiprows = 0, sheet_name = sheet_name_srpbom)
        
    if 'xlsb' in paths_cpct:
        df_cpct_busa = pd.read_excel(paths_cpct, engine='pyxlsb', skiprows = 0, sheet_name = 'BU SA')
        df_cpct_summary = pd.read_excel(paths_cpct, engine='pyxlsb', skiprows = 3, sheet_name = "Summary")
    else: 
        df_cpct_busa = pd.read_excel(paths_cpct, skiprows = 0, sheet_name = 'BU SA')
        df_cpct_summary = pd.read_excel(paths_cpct, skiprows = 3, sheet_name = "Summary")
      
        
    if 'xlsb' in sell_pricing:
        df_sell_pricing = pd.read_excel(sell_pricing, engine='pyxlsb', skiprows = 0, sheet_name = 'Master Data')
    else: 
        df_sell_pricing = pd.read_excel(sell_pricing, skiprows = 0, sheet_name = 'Master Data')

    if 'xlsb' in cost_matrix:
        df_cost_matrix = pd.read_excel(cost_matrix, engine='pyxlsb', skiprows = 1, sheet_name = "Label Matrix")
    else: 
        df_cost_matrix = pd.read_excel(cost_matrix, skiprows = 1, sheet_name = "Label Matrix")
 
    #CKIT
    if paths_ckit_1 != None:
        if 'xlsb' in paths_ckit_1:
            df_ckit_1 = pd.read_excel(paths_ckit_1, engine='pyxlsb', skiprows = 0, sheet_name = "Doc Kit SKU Summary-1")
        else:
            df_ckit_1 = pd.read_excel(paths_ckit_1, skiprows = 0, sheet_name = "Doc Kit SKU Summary-1")
        #drop NA column
        df_ckit_1 = df_ckit_1.dropna(axis=1, how = 'all')
    else:
        df_ckit_1 = pd.DataFrame(columns=list(range(10)))
        
    if paths_ckit_2 != None:
        if 'xlsb' in paths_ckit_2:
            df_ckit_2 = pd.read_excel(paths_ckit_2, engine='pyxlsb', skiprows = 0, sheet_name = "Doc Kit SKU Summary-1")
        else:
            df_ckit_2 = pd.read_excel(paths_ckit_2, skiprows = 0, sheet_name = "Doc Kit SKU Summary-1")
        df_ckit_2 = df_ckit_2.dropna(axis=1, how = 'all')
    else:
        df_ckit_2 = pd.DataFrame(columns=list(range(10)))

    ## revised for the format change on 07/07/2022 by Lily chen (lily.chen1@hp.com)
    if paths_ckit_3 != None:
        ckit_file3 = load_workbook(paths_ckit_3)
        df_ckit_3 = pd.DataFrame(columns=list(range(3)))
        for i, j in zip(ckit_file3, ckit_file3.sheetnames):
            if 'xlsb' in paths_ckit_3:
                df_ckit_3_p1 = pd.read_excel(paths_ckit_3, engine='pyxlsb', skiprows = 0, sheet_name = j)
            else:   
                df_ckit_3_p1 = pd.read_excel(paths_ckit_3, skiprows = 0, sheet_name = j)

            if j == 'Media':
                df_ckit_3_p1 = df_ckit_3_p1.iloc[:,[1,2,6]]
            else:
                df_ckit_3_p1 = df_ckit_3_p1.iloc[:,[1,2,5]]
            #df_ckit_3_p1.set_axis(df_sell_pricing.iloc[:,:3].columns,axis='columns', inplace=True)
            df_ckit_3_p1.columns = df_sell_pricing.iloc[:,:3].columns
            df_ckit_3 = pd.concat((df_ckit_3, df_ckit_3_p1), axis = 0, ignore_index=True)
            df_ckit_3 = df_ckit_3.dropna(axis=1, how = 'all')
    else:
        df_ckit_3 = pd.DataFrame(columns=list(range(3)))  
            
    df_kbd = pd.read_excel(kbd, skiprows = 0, sheet_name = "KBD")                     
    validated_date = base.range('F8').value
    month = validated_date.month
    year = validated_date.year
    
    wb = xw.Book(paths_quote)
    
    sheet_change_log = wb.sheets['Change Log']    
    platform_name = sheet_change_log.range('B2').value
    
    #if add new platform, the below code should be revised
    platform_name = platform_name.strip().rstrip('1.0').rstrip('kv').rstrip('KV').rstrip('14"').rstrip('14"').replace('BonjoviG7', 'Bonjovi').replace('DaftpunkG7', 'Daftpunk').replace('Ionian 13"', 'Ionian 13').replace('Dusker IK', 'DuskerIK').replace('IonianW 15', 'IonianW15').replace('Pecan 15"', 'Pecan15').replace('Reeds 13"', 'Reeds13').replace('Tortilla 6U 4.0', 'Tortilla6U').replace('BelugaG7', 'Beluga').strip()
    
    if platform_name == 'Pecan':
        platform_name = platform_name.replace('Pecan', 'Pecan14')
    
    df_kbd = df_kbd[df_kbd.iloc[:, 1] == platform_name ]
    df_kbd = df_kbd[df_kbd.iloc[:, 3] == int(year)]
    df_kbd = df_kbd[df_kbd.iloc[:, 4] == int(month)].iloc[:,:3]
    #delete all NA column
    df_quote_programmatrix = df_quote_programmatrix.dropna(axis=1, how = 'all')

    df_cpct_busa = df_cpct_busa.dropna(axis=1, how = 'all')
    df_sell_pricing = df_sell_pricing.dropna(axis=1, how = 'all')
    

    #copy CKIT to WILLIAM-SELL PRICING
    tmp = df_ckit_1.iloc[:,[1,2,9]]
    #tmp.set_axis(df_sell_pricing.iloc[:,:3].columns,axis='columns', inplace=True)
    df_sell_pricing.columns = df_sell_pricing.iloc[:,:3].columns
    df_sell_pricing_new = pd.concat([df_sell_pricing.iloc[:,:3], tmp])

    tmp = df_ckit_2.iloc[:,[1,2,9]]
    #tmp.set_axis(df_sell_pricing.iloc[:,:3].columns,axis='columns', inplace=True)
    df_sell_pricing.columns = df_sell_pricing.iloc[:,:3].columns
    df_sell_pricing_new = pd.concat([df_sell_pricing_new.iloc[:,:3], tmp])

    tmp = df_ckit_3.copy()
   #tmp.set_axis(df_sell_pricing.iloc[:,:3].columns,axis='columns', inplace=True)
    df_sell_pricing.columns = df_sell_pricing.iloc[:,:3].columns
    df_sell_pricing_new = pd.concat([df_sell_pricing_new.iloc[:,:3], tmp])

    
    #copy cost matrix to sell prcing
    tmp2 = df_cost_matrix[['HP P/N', 'Description', 'Approved Cost']]
    #tmp2.set_axis(df_sell_pricing.iloc[:,:3].columns,axis='columns', inplace=True)
    df_sell_pricing.columns = df_sell_pricing.iloc[:,:3].columns   
    
    #add EOLQ, DUMMY
    to_append = ['EOLQ','',0.95]
    a_series = pd.Series(to_append, index = df_sell_pricing.iloc[:,:3].columns)
    df_a_series = pd.DataFrame([a_series],columns=tmp2.columns)

    file_name_upper = re.split('\\\\|/',paths_quote)[-1].upper()
    if 'FINAL' in file_name_upper:
        to_append2 = ['Dummy','',0.00]  
    else:
        to_append2 = ['Dummy','',0.01]
        
    a_series2 = pd.Series(to_append2, index = df_sell_pricing.iloc[:,:3].columns)
    df_a_series2 = pd.DataFrame([a_series2],columns=tmp2.columns)
    
    #tmp2 = tmp2.append(a_series, ignore_index=True)
    #tmp2 = tmp2.append(a_series2, ignore_index=True)
    tmp2=pd.concat([tmp2,df_a_series], ignore_index=True)
    tmp2=pd.concat([tmp2,df_a_series2], ignore_index=True)
    df_sell_pricing_new = pd.concat([df_sell_pricing_new, tmp2])
   
    
    
    #copy KBD to sell prcing 
    tmp3 = df_kbd.iloc[:, :3]
    #tmp3.set_axis(df_sell_pricing.iloc[:,:3].columns,axis='columns', inplace=True)
    df_sell_pricing.columns = df_sell_pricing.iloc[:,:3].columns
    df_sell_pricing_new = pd.concat([df_sell_pricing_new, tmp3])

    #---------------------------------------------------------------------------------------------------------------------------- 1.	QUOTE- CKIT←→CKIT

    #CKIT
    
    #Copy busa to sell pricing
    tmp = df_cpct_busa.iloc[:,:3]
    #tmp.set_axis(df_sell_pricing_new.columns,axis='columns', inplace=True)
    df_sell_pricing.columns = df_sell_pricing.iloc[:,:3].columns
    df_busa_sellpricing = pd.concat([df_sell_pricing_new, tmp])

    #clean op
    
    df_cpct_summary_op = df_cpct_summary[df_cpct_summary.iloc[:,2].str.contains('op', na=False) | df_cpct_summary.iloc[:,2].str.contains('Op', na=False) | df_cpct_summary.iloc[:,2].str.contains('OP', na=False)| df_cpct_summary.iloc[:,2].str.contains('IMR', na=False) | df_cpct_summary.iloc[:,2].str.contains('ME', na=False)| df_cpct_summary.iloc[:,2].str.contains('IC', na=False)]
    op = []
    for i, j in zip(df_cpct_summary_op.iloc[:, 6], df_cpct_summary_op.iloc[:, 10]):
        if np.isnan(i) == True:
            op.append(j)
        else:
            op.append(i)
            
    df_cpct_summary_op['op'] = op
    df_cpct_summary_op = df_cpct_summary_op.iloc[:, [2,1,-1]]

    #copy op to sell pricing
    #df_cpct_summary_op.set_axis(df_sell_pricing.iloc[:,:3].columns,axis='columns', inplace=True)
    df_sell_pricing.columns = df_sell_pricing.iloc[:,:3].columns
    df_all_pricing = pd.concat([df_busa_sellpricing, df_cpct_summary_op])
    df_all_pricing = df_all_pricing.reset_index(drop=True)  
    
    # add one sheet named tmp
    app = xw.apps.active 
    #app.display_alerts = False
    #app.visible = False
    if 'tmp' in sheet_names: 
        wb.sheets['tmp'].delete()
    wb.sheets.add('tmp')
    sheet_tmp = wb.sheets['tmp']
    sheet_tmp.range('A1').value = df_all_pricing
    
    sheet_ckit = wb.sheets['CKIT']
    
    #CKIT USE VLOOKUP to compare price
    max_row = sheet_ckit.range('A1').expand('table').last_cell.row
    for i in range(0,max_row-1):
        sheet_ckit.range(i+2,5).value = f"=IFERROR(VLOOKUP(B{i+2},tmp!B:W,3,0), \"Null\" )"
        sheet_ckit.range(i+2,6).value = f"=IFERROR(ROUND(D{i+2}-E{i+2},2), \"Null\" )"
        if  sheet_ckit.range(i+2,6).value != 0.00:
            for j in range(1, 7):
                sheet_ckit.range(i+2,j).color = (255, 100, 255)
                


     
    #---------------------------------------------------------------------------------------------------------------MASTER DATA

# 2.	QUOTE- MASTER DATA←→ SELL PRICING(WILLIAM)-Master    
    
    sheet_masterdata = wb.sheets[sheet_name_masterdata]
    #MASTER DATA USE VLOOKUP to compare price
    max_row = sheet_masterdata.range('A1').expand('table').last_cell.row
    for i in range(0,max_row-1):
        pn = f"=IFERROR(VLOOKUP(B{i+2},tmp!B1:W40000,3,0),"
        category = f"IFERROR(VLOOKUP(C{i+2},tmp!B:W,3,0),"
        null = "\"Null\" "       
        sheet_masterdata.range(i+2,5).value = pn + category + null + "))"
        sheet_masterdata.range(i+2,6).value = f"=IFERROR(ROUND(D{i+2}-E{i+2},2)," + null + ")" 
        if sheet_masterdata.range(i+2,6).value != None:
            if sheet_masterdata.range(i+2,6).value != 0.00:
                for j in range(1, 7):
                    sheet_masterdata.range(i+2,j).color = (255, 100, 255)
                
                
        
    #----------------------------------------------------------------------------------------------------------------------------- BUSA
# 3.	QUOTE- BUSA←→CPCT-BUSA  

    sheet_busa = wb.sheets[sheet_name_busa]
    #BUSA USE VLOOKUP to compare price
    max_row = sheet_busa.range('A1').expand('table').last_cell.row
    for i in range(0,max_row-1):
        sheet_busa.range(i+2,10).value = f"=IFERROR(VLOOKUP(A{i+2},tmp!B:W,3,0), \"Null\" )"
        sheet_busa.range(i+2,11).value = f"=IFERROR(ROUND(C{i+2}-J{i+2},2), \"Null\" )"
        if sheet_busa.range(i+2,11).value != None:
            if sheet_busa.range(i+2,11).value != 0.00:
                for j in range(1, 12):
                    sheet_busa.range(i+2,j).color = (255, 100, 255)
                
               
   

        #----------------------------------------------------------------------------------------------------------------------------- PROGRAM MATRIX
#------------------------------------------ 4.	QUOTE- PROGRAM MATRIX←→SELL PRICING(WILLIAM)-Master & CPC-Option Features & CPC-BUSA
    # # add componet price
            
    sheet_program_matrix = wb.sheets[sheet_name_program_matrix]
    max_row = df_quote_programmatrix.shape[0]
    
    
    for i in range(2,max_row):
        if sheet_program_matrix.range(i+6,10).value == 'Dummy':      
            if 'FINAL' in file_name_upper:
                sheet_program_matrix.range(i+6,16).value = 0.00
            else:
                sheet_program_matrix.range(i+6,16).value = 0.01
        else: 
            sa = f"IFERROR(VLOOKUP(E{i+6},tmp!B:W,3,0)*I{i+6},"
            componets = f"IFERROR(VLOOKUP(F{i+6},tmp!B:W,3,0)*I{i+6},"
            data_source = f"IFERROR(VLOOKUP(J{i+6},tmp!B:W,3,0)*I{i+6},"
            kbd = f"IFERROR(VLOOKUP(left(E{i+6},6),tmp!B:W,3,0)*I{i+6}," 
            eolq =  f"IFERROR(VLOOKUP(c{i+6},tmp!B:W,3,0)*I{i+6},\"Null\" )"
            exist = f"=IF(LEN(TRIM(K{i+6}))=0,0,"
            sheet_program_matrix.range(i+6,16).value = exist+sa+componets+data_source+kbd+eolq+")))))"
            

        
    #add sum of price
    av_index = df_quote_programmatrix[df_quote_programmatrix.loc[:,'AV Level 2'].isnull() == False].index.tolist()
    av_index = df_quote_programmatrix[df_quote_programmatrix.loc[:,'AV Level 2'].isnull() == False].index.tolist()

    av_index_end = av_index.copy()
    av_index_end.pop(0)
    av_index_end.insert(len(av_index_end)+1, av_index[-1]+100)
    for i, j in zip(av_index, av_index_end):
        sheet_program_matrix.range(i+6,15).value = f"=SUM(P{i+6}:P{j+5})"
        sheet_program_matrix.range(i+6,18).value = f"=IFERROR(ROUND(M{i+6}-O{i+6},2), \"Null\")"
        if sheet_program_matrix.range(i+6,13).value != None and sheet_program_matrix.range(i+6,18).value != 0.00:
                if sheet_program_matrix.range(i+6,13).value != "NA" and sheet_program_matrix.range(i+6,13).value != "NotAvaible" and sheet_program_matrix.range(i+6,13).value != "#N/A" and sheet_program_matrix.range(i+6,13).value != None:
                    for k in range(0, 16): #column
                        sheet_program_matrix.range(i+6,k+1).color = (255, 100, 255)
        
        

                        
        #--------------------------------------------------------------------------------------------------        
#Program Matrix vs. AV Summary
    sheet_av_summary = wb.sheets[sheet_name_avsummary]
    max_row = sheet_av_summary.range('A1').expand('table').last_cell.row
    for i in range(0,max_row-1):
        sheet_av_summary.range(i+2,5).value = f"=IFERROR(VLOOKUP(C{i+2},'{sheet_name_program_matrix}'!D:M,10,0), \"Null\")"
        if sheet_av_summary.range(i+2,5).value != None:
            sheet_av_summary.range(i+2,5).value = f"=IFERROR(round(VLOOKUP(C{i+2},'{sheet_name_program_matrix}'!D:M,10,0),2), \"Null\")"
        sheet_av_summary.range(i+2,6).value = f"=IFERROR(ROUND(D{i+2}-E{i+2},2), \"Null\")"
        if sheet_av_summary.range(i+2,6).value != None:
            if sheet_av_summary.range(i+2,6).value != 0.00:
                for j in range(1, 7):
                    sheet_av_summary.range(i+2,j).color = (255, 100, 255)
                
                

        #--------------------------------------------------------------------------------------------------        
#SKU Summary vs. SRP BOM
    sheet_srp_bom = wb.sheets[sheet_name_srpbom]
    srp_bom_max_row = sheet_srp_bom.range('A1').expand('table').last_cell.row
    srp_bom_max_column = sheet_srp_bom.range('A1').expand('table').last_cell.column
    
    letter = []
    letter2 = []
    if srp_bom_max_column == 4:
        letter = 'D'
        letter2 = 'E'
    elif srp_bom_max_column == 3:
        letter = 'C'
        letter2 = 'D'
    elif srp_bom_max_column == 5:
        letter = 'E'
        letter2 = 'F'    
    elif srp_bom_max_column == 6:
        letter = 'F' 
        letter2 = 'G'        
        
    for i in range(0,srp_bom_max_row-1):
        sheet_srp_bom.range(i+2,srp_bom_max_column+1 ).value = f"=IF(A{i+2}=A1,,SUMIF(A:A,A{i+2},{letter}:{letter}))"

    sheet_sku_summary = wb.sheets[sheet_name_skusummary]
    sku_summary_max_row = sheet_sku_summary.range('A1').expand('table').last_cell.row
    sku_summary_max_column = sheet_sku_summary.range('A1').expand('table').last_cell.column
    for i in range(0,sku_summary_max_row-1):
        sheet_sku_summary.range(i+2,4).value = f"=IFERROR(VLOOKUP(A{i+2},'{sheet_name_srpbom}'!A:D,4,0), \"Null\")"
        if sheet_sku_summary.range(i+2,4).value != None:
            sheet_sku_summary.range(i+2,4).value = f"=IFERROR(round(VLOOKUP(A{i+2},'{sheet_name_srpbom}'!A:{letter2},{srp_bom_max_column+1},0),2), \"Null\")"
            sheet_sku_summary.range(i+2,5).value = f"=ROUND(B{i+2}-D{i+2},2)"
        if sheet_sku_summary.range(i+2,5).value != None:
            if sheet_sku_summary.range(i+2,5).value != 0.00:
                for j in range(1, 6):
                    sheet_sku_summary.range(i+2,j).color = (255, 100, 255)
                
       
        
    #Program Matrix has repetitive AV? 
    repetition1_index = ''
    tmp = df_quote_programmatrix.loc[:,'AV Level 2'].dropna()
    df1_tmp = df_quote_programmatrix[df_quote_programmatrix['AV Level 2'].isin(tmp)]
    if len(df1_tmp[df1_tmp.duplicated(subset='AV Level 2')]) != 0: 
        repetition1 = df1_tmp[df1_tmp.duplicated(subset='AV Level 2') & df1_tmp['Total Unit Price'] != 0].loc[:,['AV Level 2', 'Total Unit Price']]
        repetition1_index = repetition1.index.tolist()
        
        
    #show in the sheet
    if len(repetition1_index) != 0:
        sheet_program_matrix.range('T1').value = 'Data has repetition AV as below'       
        sheet_program_matrix.range('A1:X1').color = (102, 255, 178)
        sheet_program_matrix.range('T2').value = repetition1  
    #pop up alert
        program_matirx_warn_code()
    
    #SRP BOM has repetitive AV?
    repetition2_index = ''
    tmp = df_quote_srpbom.iloc[:,1].dropna()
    df2_tmp = df_quote_srpbom[df_quote_srpbom.iloc[:,1].isin(tmp)]
    if len(df2_tmp[df2_tmp.duplicated(subset=df_quote_srpbom.columns[:1])]) != 0:
        repetition2 = df2_tmp[df2_tmp.duplicated(subset=df_quote_srpbom.columns[:2])]
        repetition2_index = repetition2.index.tolist()
        #print(repetition2) 
        
    #show in the sheet
    if len(repetition2_index) != 0:
        sheet_srp_bom.range('I1').value = 'Data has repetition SKU+AV as below'       
        sheet_srp_bom.range('I1:L1').color = (102, 255, 178)
        sheet_srp_bom.range('I2').value = repetition2
    #pop up alert
        srp_warn_code()

        
        
        
        

#------------------------------------------------------------------------------------------------------------------------------------------
    #find Prices are different in the Combined data?
    tmp = df_all_pricing.iloc[:,0].dropna()
    df_tmp = df_all_pricing[df_all_pricing.iloc[:,0].isin(tmp)]
    repetition_index = []
    if len(df_tmp[df_tmp.duplicated()]) != 0:
        non_repetition = df_tmp[~df_tmp.duplicated(subset=df_tmp.columns[[0,2]].tolist())]
    else:
        non_repetition = df_tmp
    repetition = non_repetition[non_repetition.duplicated(subset=non_repetition.columns[:1].tolist())]
    repetition_pn = repetition.iloc[:,0].values.tolist()
    df_repetition = non_repetition[non_repetition.iloc[:,0].isin(repetition_pn)]
    repetition_index = non_repetition.index.tolist()  
    #color
    if len(repetition_index) != 0:
        sheet_tmp.range('F1').value = 'Data has repetition PN with different price as below'       
        sheet_tmp.range('F1:J1').color = (102, 255, 178)
        sheet_tmp.range('F2').value = df_repetition   
    
        
    # Create target Directory if don't exist
    dirName = '\\'.join(re.split('\\\\|/', paths_quote)[:-1])+'_revised'
    
    if not os.path.exists(dirName):
        os.mkdir(dirName)     
    wb.save(dirName + '/' + re.split('\\\\|/', paths_quote)[-1])
    #wb.save(dirName + '\\' + re.split('\\\\|/', paths_quote)[-1])
    #wb.close()
    #app.kill()
        
    finish_code()

## 2-1.function2-1: Let user user know Program Matrix has repetitive AV (pop up warning buttion.) ---Start-----------------------------------------               
## PROGRAM_MATRIX_Warn_Form
class PROGRAM_MATRIX_Warn_Form(QWidget):
    def __init__(self, name = 'PROGRAM_MATRIX_Warn_Form'):
        super(PROGRAM_MATRIX_Warn_Form,self).__init__()
        self.setWindowTitle(name)
        self.resize(300,150)   # set the pop up widget's size

        # btn 1
        self.btn_done = QPushButton(self)  
        self.btn_done.setObjectName("btn_tbd_warn")  
        self.btn_done.setText("Please check your Program Matrix: Program Matrix has repetitive AV")
        
        # set the widget's layout
        layout = QVBoxLayout()
        layout.addWidget(self.btn_done)


        self.setLayout(layout)


        # set the widget's signal
        self.btn_done.clicked.connect(self.close)
        
def program_matirx_warn_code():
    app = QApplication(sys.argv)
    program_matirx_Warn_Form = PROGRAM_MATRIX_Warn_Form('Quote Validation')
    program_matirx_Warn_Form.show()
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...')
## 2-1.function2-1: Let user user know Program Matrix has repetitive AV (pop up warning buttion.) ---End-----------------------------------------               

## 2-2.function2-2: Let user user know SRP BOM has repetition SKU+AV (pop up warning buttion.) ---Start-----------------------------------------               

##SRP_Warn_Form
class SRP_Warn_Form(QWidget):
    def __init__(self, name = 'SRP_Warn_Form'):
        super(SRP_Warn_Form,self).__init__()
        self.setWindowTitle(name)
        self.resize(300,150)   # set the pop up widget's size

        # btn 1
        self.btn_done = QPushButton(self)  
        self.btn_done.setObjectName("btn_tbd_warn")  
        self.btn_done.setText("Please check your SRP BOM: SRP BOM has repetition SKU+AV")
        
        # set the widget's layout
        layout = QVBoxLayout()
        layout.addWidget(self.btn_done)


        self.setLayout(layout)


        # set the widget's signal
        self.btn_done.clicked.connect(self.close)
        
def srp_warn_code():
    app = QApplication(sys.argv)
    srp_Warn_Form = SRP_Warn_Form('Quote Validation')
    srp_Warn_Form.show()
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...')
## 2-2.function2-2: Let user user know SRP BOM has repetition SKU+AV (pop up warning buttion.) ---End-----------------------------------------               


## 2.function2: Quote Validation ---End-----------------------------------------------------------------------------------------------------------------------------------------------------------------

 
## 3.function3: Let user choose file's path(UI)---Start-----------------------------------------------------------------------------------------------------------------------------------------------------------------
# sell prcing file, ckit...
import sys
import os
from PyQt5.QtWidgets import *

class MainForm(QWidget):
    def __init__(self, name = 'MainForm'):
        super(MainForm,self).__init__()
        self.setWindowTitle(name)
        self.resize(400,300)   # set the pop up widget's size
        
        # btn 2
        self.btn_chooseFile1 = QPushButton(self)  
        self.btn_chooseFile1.setObjectName("btn_chooseFile")  
        self.btn_chooseFile1.setText("Choose CKIT NB file's Path")        
        
        # btn 3
        self.btn_chooseFile2 = QPushButton(self)  
        self.btn_chooseFile2.setObjectName("btn_chooseFile")  
        self.btn_chooseFile2.setText("Choose CKIT AIODTVRTC file's Path")

        # btn 4
        self.btn_chooseFile3 = QPushButton(self)  
        self.btn_chooseFile3.setObjectName("btn_chooseFile")  
        self.btn_chooseFile3.setText("Choose CKIT Media file's Path")

        # btn 5
        self.btn_chooseFile4 = QPushButton(self)  
        self.btn_chooseFile4.setObjectName("btn_chooseFile")  
        self.btn_chooseFile4.setText("Choose Sell_Pricing file's Path")
        
        # btn 6
        self.btn_chooseFile5 = QPushButton(self)  
        self.btn_chooseFile5.setObjectName("btn_chooseFile")  
        self.btn_chooseFile5.setText("Choose Cost_Matrix file's Path")

        # btn 7
        self.btn_chooseFile6 = QPushButton(self)  
        self.btn_chooseFile6.setObjectName("btn_chooseFile")  
        self.btn_chooseFile6.setText("Choose KBD_Price file's Path")
        
        # set the widget's layout
        layout = QVBoxLayout()
        layout.addWidget(self.btn_chooseFile1)        
        layout.addWidget(self.btn_chooseFile2)        
        layout.addWidget(self.btn_chooseFile3)
        layout.addWidget(self.btn_chooseFile4)
        layout.addWidget(self.btn_chooseFile5)
        layout.addWidget(self.btn_chooseFile6)


        self.setLayout(layout)


        # set the widget's signal
        self.btn_chooseFile1.clicked.connect(self.slot_btn_chooseFile1)
        self.btn_chooseFile2.clicked.connect(self.slot_btn_chooseFile2)
        self.btn_chooseFile3.clicked.connect(self.slot_btn_chooseFile3)
        self.btn_chooseFile4.clicked.connect(self.slot_btn_chooseFile4)
        self.btn_chooseFile5.clicked.connect(self.slot_btn_chooseFile5)
        self.btn_chooseFile6.clicked.connect(self.slot_btn_chooseFile6)
        
    def slot_btn_chooseFile1(self):
        fileName_choose, filetype = QFileDialog.getOpenFileName(self,  
                                    "Choose CKIT NB file's Path",  
                                    "", # start path
                                    "Excel File (*.xlsx *.xls *.xlsb);;All Files (*)")  

        if fileName_choose == "":
            return

#         print("\nQuote folder's Path you chose is:")
#         print(fileName_choose)
        base = xw.Book.caller().sheets['Main']
        base.range('G21').value = fileName_choose

    def slot_btn_chooseFile2(self):
        fileName_choose, filetype = QFileDialog.getOpenFileName(self,  
                                    "Choose CKIT AIODTVRTC file's Path",  
                                    "", # start path
                                    "Excel File (*.xlsx *.xls *.xlsb);;All Files (*)")  

        if fileName_choose == "":
            return

#         print("\nQuote folder's Path you chose is:")
#         print(fileName_choose)
        base = xw.Book.caller().sheets['Main']
        base.range('G22').value = fileName_choose         

    def slot_btn_chooseFile3(self):
        fileName_choose, filetype = QFileDialog.getOpenFileName(self,  
                                    "Choose CKIT Media file's Path",  
                                    "", # start path
                                    "Excel File (*.xlsx *.xls *.xlsb);;All Files (*)")  

        if fileName_choose == "":
            return

#         print("\nQuote folder's Path you chose is:")
#         print(fileName_choose)
        base = xw.Book.caller().sheets['Main']
        base.range('G23').value = fileName_choose

    def slot_btn_chooseFile4(self):
        fileName_choose, filetype = QFileDialog.getOpenFileName(self,  
                                    "Choose Sell_Pricing file's Path",  
                                    "", # start path
                                    "Excel File (*.xlsx *.xls *.xlsb);;All Files (*)")   

        if fileName_choose == "":
            return

        base = xw.Book.caller().sheets['Main']
        base.range('F25').value = fileName_choose

    def slot_btn_chooseFile5(self):
        fileName_choose, filetype = QFileDialog.getOpenFileName(self,  
                                    "Choose Cost_Matrix file's Path",  
                                    "", # start path
                                    "Excel File (*.xlsx *.xls *.xlsb);;All Files (*)")   

        if fileName_choose == "":
            return

#         print("\nQuote folder's Path you chose is:")
#         print(fileName_choose)
        base = xw.Book.caller().sheets['Main']
        base.range('F29').value = fileName_choose
        
    def slot_btn_chooseFile6(self):
        fileName_choose, filetype = QFileDialog.getOpenFileName(self,  
                                    "Choose KBD_Price file's Path",  
                                    "",#self.cwd, # start path
                                    "Excel File (*.xlsx *.xls *.xlsb);;All Files (*)")   

        if fileName_choose == "":
            return


        base = xw.Book.caller().sheets['Main']
        base.range('F33').value = fileName_choose
 
def choose_files():
    app = QApplication(sys.argv)
    mainForm = MainForm('Quote Validation')
    mainForm.show()
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...')
 ## 3.function3: Let user choose file's path(UI)---End-----------------------------------------------------------------------------------------------------------------------------------------------------------------
        
        
## 4.function4: Let user choose file's path(UI)---Start-----------------------------------------------------------------------------------------------------------------------------------------------------------------
# quote, cpct...        
class MainForm2(QWidget):
    def __init__(self, name = 'MainForm2'):
        super(MainForm2,self).__init__()
        self.setWindowTitle(name)
        self.resize(400,150)   # set the pop up widget's size

        # btn 2
        self.btn_chooseFile1 = QPushButton(self)  
        self.btn_chooseFile1.setObjectName("btn_chooseFile")  
        self.btn_chooseFile1.setText("[---Choose Quote file's Path---]")

        # btn 3
        self.btn_chooseFile2 = QPushButton(self)  
        self.btn_chooseFile2.setObjectName("btn_chooseFile")  
        self.btn_chooseFile2.setText("[---Choose CPCT file's Path---]")
        
        # set the widget's layout
        layout = QVBoxLayout()
        layout.addWidget(self.btn_chooseFile1)
        layout.addWidget(self.btn_chooseFile2)

        self.setLayout(layout)


        # set the widget's signal
        self.btn_chooseFile1.clicked.connect(self.slot_btn_chooseFile1)
        self.btn_chooseFile2.clicked.connect(self.slot_btn_chooseFile2)
        
    def slot_btn_chooseFile1(self):
        fileName_choose, filetype = QFileDialog.getOpenFileName(self,  
                                    "Choose Quote file's Path",  
                                    "", # start path
                                    "Excel File (*.xlsx *.xls *.xlsb);;All Files (*)")   


        base = xw.Book.caller().sheets['Main']
        base.range('F13').value = fileName_choose


    def slot_btn_chooseFile2(self):
        fileName_choose, filetype = QFileDialog.getOpenFileName(self,  
                                    "Choose CPCT file's Path",  
                                    "", # start path
                                    "Excel File (*.xlsx *.xls *.xlsb);;All Files (*)")  

        if fileName_choose != "":
            self.close()
        elif fileName_choose == "":
            return 

        base = xw.Book.caller().sheets['Main']
        base.range('F17').value = fileName_choose


 
def choose_files2():
    app = QApplication(sys.argv)
    mainForm2 = MainForm2('Quote Validation')
    mainForm2.show()
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...')

class MainForm3(QWidget):
    def __init__(self, name = 'MainForm3'):
        super(MainForm3,self).__init__()
        self.setWindowTitle(name)
        self.resize(400,100)   # set the pop up widget's size

        # btn 1
        self.btn_chooseDir = QPushButton(self)  
        self.btn_chooseDir.setObjectName("btn_chooseDir")  
        self.btn_chooseDir.setText("Choose CPCT's folder")

        
        # set the widget's layout
        layout = QVBoxLayout()
        layout.addWidget(self.btn_chooseDir)


        self.setLayout(layout)


        # set the widget's signal
        self.btn_chooseDir.clicked.connect(self.slot_btn_chooseDir)




    def slot_btn_chooseDir(self):
        dir_choose = QFileDialog.getExistingDirectory(self,  
                                    "Choose CPCT's folder",  
                                    "") # start path

        if dir_choose != "":
            self.close()
        elif dir_choose == "":
            return      

        base = xw.Book.caller().sheets['Main']
        base.range('F4').value = dir_choose            
 
def choose_files3():
    app = QApplication(sys.argv)
    mainForm3 = MainForm3('Delete CPCT non-effectively date')
    mainForm3.show()
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...')



class CloseForm(QWidget):
    def __init__(self, name = 'CloseForm'):
        super(CloseForm,self).__init__()
        self.setWindowTitle(name)
        self.resize(200,100)   # set the pop up widget's size

        # btn 1
        self.btn_done = QPushButton(self)  
        self.btn_done.setObjectName("btn_done")  
        self.btn_done.setText("Done")
        
        # set the widget's layout
        layout = QVBoxLayout()
        layout.addWidget(self.btn_done)


        self.setLayout(layout)


        # set the widget's signal
        self.btn_done.clicked.connect(self.close)
## 4.function4: Let user choose file's path(UI)---End-----------------------------------------------------------------------------------------------------------------------------------------------------------------
       
## 5.function5: Quote Consolidation ---Start--------------------------------------------------------------------------------------------------------------------------------------------------------------------
class ConsolidationForm(QWidget):
    def __init__(self, name = 'ConsolidationForm'):
        super(ConsolidationForm,self).__init__()
        self.setWindowTitle(name)
        self.resize(400,150)   # set the pop up widget's size

        # btn 1
        self.btn_chooseFile1 = QPushButton(self)  
        self.btn_chooseFile1.setObjectName("btn_chooseFile")  
        self.btn_chooseFile1.setText("[---Choose Quote file's Path---]")
        
        # set the widget's layout
        layout = QVBoxLayout()
        layout.addWidget(self.btn_chooseFile1)

        self.setLayout(layout)

        # set the widget's signal
        self.btn_chooseFile1.clicked.connect(self.slot_btn_chooseFile1)
    
    def slot_btn_chooseFile1(self):
        global fileName_choose
        fileName_choose, filetype = QFileDialog.getOpenFileNames(self,  
                                    "Choose Quote file's Path",  
                                    "", # start path
                                    "Excel File (*.xlsx *.xls *.xlsb);;All Files (*)")   

        if fileName_choose != "":
            self.close()

def quote_consolidation():
    consolidation_getdateForm()
    app = QApplication(sys.argv)
    consolidationForm = ConsolidationForm('Quote Consolidation')
    consolidationForm.show()
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...')
    
    new_effective_day = 'Effective '+ text
    df = pd.DataFrame(columns=['AV', 'Description', new_effective_day, 'Platform', 'ODM', 'Quote File Name'])
    df2 = pd.DataFrame(columns=['SKU', 'Description', new_effective_day, 'Platform', 'ODM', 'Quote File Name'])
    test_platform = []
    for paths_quote in fileName_choose:
        
        test_platform.append(paths_quote)
        #get quote_program_matrix sheet's name
        if 'xlsb' in paths_quote:
            quote_program_matrix = pd.ExcelFile(paths_quote, engine='pyxlsb')
        else:
            quote_program_matrix = pd.ExcelFile(paths_quote)

        sheet_names = quote_program_matrix.sheet_names 

        sheet_name_change_log = []
        sheet_name_avsummary = []
        sheet_name_skusummary = []
        for i in sheet_names:
            if 'Change Log' in i:
                sheet_name_changelog = i
            if 'Change log' in i:
                sheet_name_changelog = i
            if 'AV Summary' in i: 
                sheet_name_avsummary = i
            if 'AV summary' in i: 
                sheet_name_avsummary = i
            if 'SKU Summary' in i: 
                sheet_name_skusummary = i
            if 'SKU summary' in i: 
                sheet_name_skusummary = i

        #get platform & ODM's name
        import xlwings as xw
        app = xw.App(visible=False, add_book=False) # can't see the file
        app.display_alerts = False # close alert
        app.screen_updating = False # close screen update
        wb = app.books.open(paths_quote, update_links=False, read_only=True, ignore_read_only_recommended=True)
        
        sheet_change_log = wb.sheets[sheet_name_changelog]    
        platform_name = sheet_change_log.range('B2').value
        odm_name = sheet_change_log.range('B3').value
        quote_file_name = paths_quote.split("/")[-1]

        #load the data
        if 'xlsb' in paths_quote:
            df_quote_avsummary = pd.read_excel(paths_quote, engine='pyxlsb', skiprows = 0, sheet_name = sheet_name_avsummary)
            df_quote_skusummary = pd.read_excel(paths_quote, engine='pyxlsb', skiprows = 0, sheet_name = sheet_name_skusummary)
        else: 
            df_quote_avsummary = pd.read_excel(paths_quote, skiprows = 0, sheet_name = sheet_name_avsummary)
            df_quote_skusummary = pd.read_excel(paths_quote, skiprows = 0, sheet_name = sheet_name_skusummary)

        # get "AV Summary" & "SKU Summary" columns' name
        df_quote_avsummary_column_name = df_quote_avsummary.columns.tolist()
        for i in df_quote_avsummary_column_name:
            if 'Effective' in i:
                effective_column_name = i
        df_quote_avsummary = df_quote_avsummary.reindex(columns=['AV', 'Description', effective_column_name])

        df_quote_skusummary_column_name = df_quote_skusummary.columns.tolist()
        for i in df_quote_skusummary_column_name:
            if 'Effective' in i:
                effective_column_name2 = i
        df_quote_skusummary = df_quote_skusummary.reindex(columns=['SKU', 'Description', effective_column_name2])

        # get "AV Summary" & "SKU Summary" columns' info
        df_quote_avsummary.rename(columns={effective_column_name:new_effective_day}, inplace=True)
        df_quote_avsummary['Platform'] = platform_name
        df_quote_avsummary['ODM'] = odm_name
        df_quote_avsummary['Quote File Name'] = quote_file_name
        df_quote_avsummary = df_quote_avsummary[~df_quote_avsummary['AV'].isnull()] #delete row which AV is null

          # update comment
        df_quote_avsummary['Comment'] = ''
        df_list = []
        for i in df_quote_avsummary[new_effective_day]:
            if isinstance(i,str):
                df_list.append(np.nan)       
            else:
                df_list.append(i)
        df_quote_avsummary[new_effective_day] = df_list        
        # df_quote_avsummary.loc[df_quote_avsummary[new_effective_day] == '0x2a', new_effective_day] = np.nan
        # df_quote_avsummary.loc[df_quote_avsummary[new_effective_day] == '0xf', new_effective_day] = np.nan
        df_quote_avsummary.loc[df_quote_avsummary[new_effective_day] <= 0, 'Comment'] = 'AV Cost Not Updated'
        df_quote_avsummary.loc[df_quote_avsummary[new_effective_day].isnull(), new_effective_day] = 'NULL'
        df_quote_av_not_duplicate = df_quote_avsummary[~df_quote_avsummary.duplicated(subset=['AV',new_effective_day])]
        df_quote_av_same = df_quote_av_not_duplicate[df_quote_av_not_duplicate.duplicated(subset='AV')]['AV'].tolist()
        df_quote_avsummary.loc[df_quote_avsummary['AV'].isin(df_quote_av_same), 'Comment'] = 'Same AV With Different Cost'
        df = pd.concat([df, df_quote_avsummary], ignore_index=True) 


        df_quote_skusummary.rename(columns={effective_column_name2:new_effective_day}, inplace=True)
        df_quote_skusummary['Platform'] = platform_name
        df_quote_skusummary['ODM'] = odm_name
        df_quote_skusummary['Quote File Name'] = quote_file_name
        df_quote_skusummary = df_quote_skusummary[~df_quote_skusummary['SKU'].isnull()] #delete row which SKU is null


          # update comment
        df_quote_skusummary['Comment'] = ''
        df_list2 = []
        for i in df_quote_skusummary[new_effective_day]:
            if isinstance(i,str):
                df_list2.append(np.nan)       
            else:
                df_list2.append(i)
        df_quote_skusummary[new_effective_day] = df_list2        
        # df_quote_skusummary.loc[df_quote_skusummary[new_effective_day] == '0x2a', new_effective_day] = np.nan
        # df_quote_skusummary.loc[df_quote_skusummary[new_effective_day] == '0xf', new_effective_day] = np.nan
        df_quote_skusummary.loc[df_quote_skusummary[new_effective_day] <= 0, 'Comment'] = 'SKU Cost Not Updated'
        df_quote_skusummary.loc[df_quote_skusummary[new_effective_day].isnull(), new_effective_day] = ''
        df_quote_sku_not_duplicate = df_quote_skusummary[~df_quote_skusummary.duplicated(subset=['SKU',new_effective_day])]
        df_quote_sku_same = df_quote_sku_not_duplicate[df_quote_sku_not_duplicate.duplicated(subset='SKU')]['SKU'].tolist()
        df_quote_skusummary.loc[df_quote_skusummary['SKU'].isin(df_quote_sku_same), 'Comment'] = 'Same SKU With Different Cost'
        df2 = pd.concat([df2, df_quote_skusummary], ignore_index=True)

        wb.close() # close file
        app.quit() # close app

    #save file
        # Create target Directory if it doesn't exist
    dirName = '/'.join(fileName_choose[0].split('/')[:-2])+'/Quote_Consolidation_' + text 
    if not os.path.exists(dirName):
        os.mkdir(dirName) 

        #create a Pandas Excel writer using XlsxWriter as the engine
    writer = pd.ExcelWriter(dirName+'/Quote_Consolidation_'+text+'_'+datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')+'.xlsx', engine='xlsxwriter')

        #write each DataFrame to a specific sheet
    df.to_excel(writer, sheet_name='AV Summary', index=False)
    df2.to_excel(writer, sheet_name='SKU Summary', index=False)

    #format header
        # Get the xlsxwriter workbook and worksheet objects.
    workbook  = writer.book
    worksheet = writer.sheets['AV Summary']
    worksheet2 = writer.sheets['SKU Summary']

        # Add a header format.
    header_format = workbook.add_format({
        'bold': True,
        'text_wrap': True,
        'valign': 'top',
        'fg_color': '#8EA9DB'})

    worksheet.set_column("A:G", 25)
    worksheet2.set_column("A:G", 25)

        # Write the column headers with the defined format.
    for col_num, value in enumerate(df.columns.values):
        worksheet.write(0, col_num, value, header_format)
    for col_num, value in enumerate(df2.columns.values):
        worksheet2.write(0, col_num, value, header_format)


        #close the Pandas Excel writer and output the Excel file
    #writer.save()
    writer.close()
    #print save file's path
    base = xw.Book.caller().sheets['Quote Consolidation'] 
    base.range('B19').value = dirName+'/Quote_Consolidation_'+text+'_'+datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')+'.xlsx'
    df3 = pd.DataFrame(test_platform)
    base.range('C29', 'O311').clear_contents()
    base.range('C29').value = df3
    finish_code()

## 5-1.function5-1: Let user enter effectively date ---Start------------------------------------------------------------------
class GetdateForm(QWidget):
    def __init__(self, name = 'GetdateForm'):
        super(GetdateForm,self).__init__()
        self.setWindowTitle(name)
        self.resize(400,150)   # set the pop up widget's size
        # btn 1
        self.btn_chooseFile1 = QPushButton(self)  
        self.btn_chooseFile1.setObjectName("btn_chooseFile")  
        self.btn_chooseFile1.setText("[---Enter Effective Month---]")
        
        # set the widget's layout
        layout = QVBoxLayout()
        layout.addWidget(self.btn_chooseFile1)

        self.setLayout(layout)

        # set the widget's signal
        self.btn_chooseFile1.clicked.connect(self.slot_btn_chooseFile1)
        self.close()
    def slot_btn_chooseFile1(self):
        global text
        text, okPressed = QInputDialog.getText(self, "Get text","Please enter Effective Month as in format like Sep-2021", QLineEdit.Normal, "")
        if okPressed and text != '':
            self.close()
        
def consolidation_getdateForm():
    app = QApplication(sys.argv)
    getdateForm = GetdateForm('Quote Consolidation')
    getdateForm.show()
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...')

## 5-1.function5-1: Let user enter effectively date ---End------------------------------------------------------------------


## 5.function5: Quote Consolidation ---End--------------------------------------------------------------------------------------------------------------------------------------------------------------------

           
       





 
